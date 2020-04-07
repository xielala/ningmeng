
#地址 测试数据 断言期望结果  除了这几个 其他基本高度相似  80%
#参数化 url data excepted


#存到Excel里面 python去操作excel
#1.只支持 这种后缀：。xlsx  -->openpyxl  只支持这种格式

from openpyxl import load_workbook

#1.打开excel
wb=load_workbook('test.xlsx')

#2.定位表单
sheet=wb['python']

#3.定位单元格  行列值去定位
value=sheet.cell(1,1).value
print('最大行：{0}'.format(sheet.max_row))
print('最大列：{0}'.format(sheet.max_column))
print(value)

#数据从excel里拿出来是什么类型
#数字还是数字，字符串是字符串
print('url:{}，类型是：{}'.format(sheet.cell(1,1).value,type(sheet.cell(1,1).value)))
print('data:{}，类型是：{}'.format(sheet.cell(1,2).value,type(sheet.cell(1,2).value)))
print('excepted:{}，类型是：{}'.format(sheet.cell(1,3).value,type(sheet.cell(1,3).value)))
print('method:{}，类型是：{}'.format(sheet.cell(1,4).value,type(sheet.cell(1,4).value)))


# 作业
# 1.把这些数据存在Excel里面，格式请看视频
# 2.利用openpyxl写一个专门读取Excel里面测试数据的类
# 3.结合课堂上老师讲解的单元测试方法，通过初始化函数传参的方法，完成单元测试
# 4.提交操作Excel的类、test_suit、test_http类（类名可以自己定，）