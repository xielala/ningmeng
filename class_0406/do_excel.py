
from openpyxl import load_workbook
from class_0406.read_config import ReadConfig

class DoExcel:

    def __init__(self,file_name,sheet_name):
        self.file_name=file_name
        self.sheet_name=sheet_name

    def get_data(self):
        """
        mode是控制是否执行所有用例  默认值是all 为all就执行所有用例
        如果不等于all 就进行分值判断
        button的值  只能输入all 列表  这两种类型
        """

        #从配置文件读取mode值
        mode=ReadConfig().read_config('case.config','PYTHON11','num')
        wb=load_workbook(self.file_name)
        sheet=wb[self.sheet_name]
        test_data=[]

        for i in range(2,sheet.max_row+1):#2 3 4 5
            sub_data={}
            sub_data['case_id'] = sheet.cell(i, 1).value
            sub_data['url']=sheet.cell(i,1).value
            sub_data['data'] = sheet.cell(i, 2).value
            sub_data['method'] = sheet.cell(i, 3).value
            sub_data['except'] = sheet.cell(i, 4).value
            test_data.append(sub_data)

        #根据button值去判断
        if mode=='all':#执行所有的用例
            final_data=test_data
        else:
            final_data=[]
            for item in test_data:#对test_data的所有用例进行遍历
                if item['case_id'] in eval(mode):
                    final_data.append(item)

        return final_data

if __name__ == '__main__':
    print(DoExcel('../../API_AUTO/tools/test.xlsx', 'python').get_data())


# 作业：
# 1.把列表里的字典存到excel里面，格式请看视频
# 2.利用openpyxl写一个专门读取Excel里面测试数据的类
# 3.结合课堂上老师讲解的单元测试方法，通过ddt的方法，完成单元测试
# 4.完成用例的可配置操作（不一定要按照老师的，你们可以自行设计，原理不变即可）