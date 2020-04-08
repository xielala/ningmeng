
from openpyxl import load_workbook

class DoExcel:

    def __init__(self,file_name,sheet_name):
        self.file_name=file_name
        self.sheet_name=sheet_name

    def get_header(self):
        """获取第一行的标题"""
        wb = load_workbook(self.file_name)
        sheet = wb[self.sheet_name]
        header=[] #存储我们的标题行
        for j in range(1,sheet.max_column+1):
            header.append(sheet.cell(1,j).value)
        return header

    def get_data(self):
        """根据嵌套循环读取数据"""
        wb=load_workbook(self.file_name)
        sheet=wb[self.sheet_name]
        test_data=[]
        header=self.get_header()
        for i in range(2,sheet.max_row+1):#2 3 4 5
            sub_data={}
            for j in range(1,sheet.max_column+1): #1 2 3 4
                sub_data[header[j-1]]=sheet.cell(i,j).value
            test_data.append(sub_data)

        return test_data

if __name__ == '__main__':
    print(DoExcel('../../API_AUTO/tools/test.xlsx', 'python').get_data())