
from openpyxl import load_workbook

class DoExcel:

    def __init__(self,file_name,sheet_name):
        self.file_name=file_name
        self.sheet_name=sheet_name

    def get_header(self):

        wb = load_workbook(self.file_name)
        sheet = wb[self.sheet_name]
        header=[] #存储我们的标题行
        for j in range(1,sheet.max_column+1):
            header.append(sheet.cell(1,j).value)
        return header

    def get_data(self,mode='all'):
        """
        mode是控制是否执行所有用例  默认值是all 为all就执行所有用例
        如果不等于all 就进行分值判断
        button的值  只能输入all 列表  这两种类型
        """
        wb=load_workbook(self.file_name)
        sheet=wb[self.sheet_name]
        test_data=[]

        for i in range(2,sheet.max_row+1):#2 3 4 5
            sub_data={}
            sub_data['case_id'] = sheet.cell(i, 1).value
            sub_data['url']=sheet.cell(i,1).value
            sub_data['data'] = sheet.cell(i, 2).value
            sub_data['method'] = sheet.cell(i, 3).value
            sub_data['except'] = sheet.cell(i, 4).value
            test_data.append(sub_data)

        #根据button值去判断
        if mode=='all':#执行所有的用例
            final_data=test_data
        else:
            final_data=[]
            for item in test_data:#对test_data的所有用例进行遍历
                if item['case_id'] in mode:
                    final_data.append(item)

        return final_data

if __name__ == '__main__':
    print(DoExcel('../../API_AUTO/tools/test.xlsx', 'python').get_data())